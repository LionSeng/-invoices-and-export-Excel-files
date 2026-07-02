import io, os, shutil, json, re
from datetime import datetime
from flask import Flask, render_template, request, send_file, jsonify
import openpyxl

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
HISTORY_PATH = os.path.join(BASE_DIR, "history.json")

CUSTOMERS = {
    "PureSpectrum": {
        "display_name": "PureSpectrum",
        "template_file": os.path.join(BASE_DIR, "template_PureSpectrum.xlsx"),
        "output_format": "PureSpectrum API-E-invoice#{code}.xlsx",
        "invoice_pattern": "BKM-{YY}{MM}-Pure",
        "desc_pattern": "{MonthName} API",
        "date_cell": "G4",
        "fields": [
            {"id": "invoice_no", "label": "Invoice No", "type": "text", "cell": "G5"},
            {"id": "description", "label": "Project Description", "type": "text", "cell": "G6"},
            {"id": "item_desc", "label": "Item Description", "type": "text", "cell": "C16", "inherit_from": "description"},
            {"id": "amount", "label": "Amount (USD)", "type": "number", "cell": "G16"},
        ],
    },
    "DataSpring": {
        "display_name": "DataSpring",
        "template_file": os.path.join(BASE_DIR, "template_DataSpring.xlsx"),
        "output_format": "DataSpring API E-invoice#{code}.xlsx",
        "invoice_pattern": "BKM-{YYYY}S-{MM}",
        "desc_pattern": "{MonthName} API",
        "date_cell": "G4",
        "fields": [
            {"id": "invoice_no", "label": "Invoice No", "type": "text", "cell": "G5"},
            {"id": "description", "label": "Project Description", "type": "text", "cell": "G6"},
            {"id": "item_desc", "label": "Item Description", "type": "text", "cell": "A16", "inherit_from": "description"},
            {"id": "amount", "label": "Amount (USD)", "type": "number", "cell": "F16"},
        ],
    },
}

MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]
MONTH_ABBRS = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

def resolve_pattern(pattern, year, month):
    mm = f"{month:02d}"
    m = str(month)
    yy = f"{year % 100:02d}"
    yyyy = str(year)
    mn = MONTH_NAMES[month]
    ma = MONTH_ABBRS[month]
    s = pattern.replace("{YY}", yy).replace("{YYYY}", yyyy)
    s = s.replace("{MM}", mm).replace("{M}", m)
    s = s.replace("{MonthName}", mn).replace("{MonthAbbr}", ma)
    return s

def generate_excel(customer_key, raw_fields, month_str):
    cfg = CUSTOMERS[customer_key]
    parts = month_str.split("-")
    year, month = int(parts[0]), int(parts[1])
    today_date = datetime.now()
    code = "{:04d}{:02d}".format(year, month)

    auto_no = resolve_pattern(cfg.get("invoice_pattern", ""), year, month) if cfg.get("invoice_pattern") else ""
    auto_desc = resolve_pattern(cfg.get("desc_pattern", ""), year, month) if cfg.get("desc_pattern") else ""

    resolved = {}
    for field in cfg["fields"]:
        fid = field["id"]
        raw = raw_fields.get(fid, "").strip()
        if raw:
            if field.get("type") == "number":
                try:
                    resolved[fid] = float(raw)
                except ValueError:
                    resolved[fid] = raw
            else:
                resolved[fid] = raw
        elif fid == "invoice_no" and auto_no:
            resolved[fid] = auto_no
        elif fid == "description" and auto_desc:
            resolved[fid] = auto_desc
        elif field.get("inherit_from") and resolved.get(field["inherit_from"]):
            resolved[fid] = resolved[field["inherit_from"]]

    wb = openpyxl.load_workbook(cfg["template_file"])
    ws = wb.active

    if cfg.get("date_cell"):
        ws[cfg["date_cell"]] = today_date

    for field in cfg["fields"]:
        cell_ref = field.get("cell")
        if cell_ref and field["id"] in resolved:
            ws[cell_ref] = resolved[field["id"]]

    out_name = cfg["output_format"].replace("{code}", code)
    out_name = out_name.replace("{date}", month_str)
    out_path = os.path.join(BASE_DIR, "outputs", out_name)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    wb.close()

    _record_history(
        customer_key, cfg["display_name"], month_str, code,
        resolved.get("invoice_no", ""),
        resolved.get("description", ""),
        str(resolved.get("amount", "")),
    )
    return out_path, out_name

def _record_history(customer_key, customer_name, month_str, code, invoice_no, description, amount):
    entries = []
    if os.path.exists(HISTORY_PATH):
        # Backup before modifying
        bak_path = HISTORY_PATH + ".bak"
        try:
            shutil.copy2(HISTORY_PATH, bak_path)
        except IOError:
            pass
        try:
            with open(HISTORY_PATH, "r", encoding="utf-8") as f:
                entries = json.load(f)
        except (json.JSONDecodeError, IOError):
            entries = []
    entry = {
        "customer_key": customer_key, "customer_name": customer_name,
        "month": month_str, "code": code,
        "invoice_no": invoice_no, "description": description, "amount": amount,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    entries.append(entry)
    with open(HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(entries, f, indent=2, ensure_ascii=False)

@app.route("/")
def index():
    return render_template("index.html", customers=CUSTOMERS)

@app.route("/generate", methods=["POST"])
def generate():
    customer_key = request.form.get("customer")
    month_str = request.form.get("month", datetime.now().strftime("%Y-%m"))
    if customer_key not in CUSTOMERS:
        return "Invalid customer", 400
    try:
        out_path, out_name = generate_excel(customer_key, request.form, month_str)
        return send_file(out_path, as_attachment=True, download_name=out_name,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        import traceback; traceback.print_exc()
        return f"Error generating Excel: {e}", 500

@app.route("/history")
def history():
    entries = []
    if os.path.exists(HISTORY_PATH):
        try:
            with open(HISTORY_PATH, "r", encoding="utf-8") as f:
                entries = json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return jsonify(entries=entries)

if __name__ == "__main__":
    port = 5000
    print(f"  Invoice Generator started -> http://localhost:{port}")
    print("  Press Ctrl+C to stop.")
    app.run(host="127.0.0.1", port=port, debug=False)
